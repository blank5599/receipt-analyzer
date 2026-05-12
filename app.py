import os
import io
import json
import base64
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file
from dotenv import load_dotenv
from google import genai
from google.genai import types as genai_types
import PIL.Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import init_db, get_conn, next_voucher_no

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
if not GEMINI_API_KEY:
    raise RuntimeError(".env 파일에 GEMINI_API_KEY가 설정되지 않았습니다.")

client = genai.Client(api_key=GEMINI_API_KEY)
MODEL = "gemini-3.1-flash-lite"

app = Flask(__name__)
init_db()

PROMPT = """\
이 영수증 이미지를 분석하여 아래 항목을 JSON 형식으로 추출해주세요.

추출 항목:
- store_name (string): 가맹점명 또는 상호명
- payment_datetime (string): 결제 일시를 "YYYY-MM-DD HH:MM" 형식으로. 없으면 null
- total_amount (number): 최종 결제 금액(원). 쉼표·원 기호 제거 후 숫자만. 없으면 null
- items (array): 개별 품목 목록. 각 항목은 다음 필드를 포함:
    - name (string): 품목명 (한글 그대로 유지)
    - quantity (number): 수량. 명시되지 않으면 1
    - unit_price (number): 단가(원). 쉼표·원 기호 제거 후 숫자만. 없으면 null

규칙:
1. 반드시 유효한 JSON만 반환하세요. 마크다운 코드블록(```)이나 설명 텍스트 없이 순수 JSON만 출력하세요.
2. 할인, 부가세, 봉사료 등도 items에 포함하되 품목명을 명확히 표시하세요.
3. 인식할 수 없거나 없는 항목은 null로 처리하세요.
"""


# ── Excel 스타일 헬퍼 ──────────────────────────────────────────────────────

def _font(bold=False, color="1C1C1C", size=10):
    return Font(name="맑은 고딕", bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin", color="D0D8E8")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_row(ws, row, values, bold=False, bg=None, fg="1C1C1C", size=10, aligns=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=bold, color=fg, size=size)
        c.border = _thin_border()
        ah = (aligns[col - 1] if aligns and col <= len(aligns) else "center")
        c.alignment = _align(h=ah)
        if bg:
            c.fill = _fill(bg)


def build_excel_one(r, items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "전표상세"

    # 열 너비
    for col, w in zip("ABCDE", [16, 30, 10, 14, 14]):
        ws.column_dimensions[col].width = w

    # ── 타이틀 ──
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "한국도로공사  ·  지출 전표"
    c.font = _font(bold=True, color="FFFFFF", size=14)
    c.fill = _fill("003087")
    c.alignment = _align()
    ws.row_dimensions[1].height = 38

    # ── 전표 정보 ──
    meta = [
        ("전표번호",   r["voucher_no"]),
        ("가맹점명",   r["store_name"] or "-"),
        ("결제일시",   r["payment_datetime"] or "-"),
        ("비용분류",   r["category"] or "-"),
        ("총 결제금액", f"{int(r['total_amount']):,}원" if r["total_amount"] else "-"),
        ("처리일시",   r["created_at"]),
    ]
    for i, (label, val) in enumerate(meta, 2):
        ws.row_dimensions[i].height = 22
        # 레이블 셀
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _font(bold=True, color="003087", size=10)
        lc.fill = _fill("E4EDFA")
        lc.alignment = _align()
        lc.border = _thin_border()
        # 값 셀 (B~E 병합)
        ws.merge_cells(f"B{i}:E{i}")
        vc = ws.cell(row=i, column=2, value=val)
        vc.font = _font(bold=(label == "총 결제금액"), color="E8112D" if label == "총 결제금액" else "1C1C1C", size=10)
        vc.alignment = _align(h="left")
        vc.border = _thin_border()

    # ── 품목 테이블 ──
    gap = len(meta) + 3
    ws.row_dimensions[gap].height = 24
    headers = ["#", "품목명", "수량", "단가 (원)", "금액 (원)"]
    aligns  = ["center", "left", "center", "right", "right"]
    _set_row(ws, gap, headers, bold=True, bg="003087", fg="FFFFFF", size=10, aligns=aligns)

    for idx, item in enumerate(items, 1):
        r_num = gap + idx
        qty   = item.get("quantity") or 1
        up    = item.get("unit_price")
        total = qty * up if up is not None else None
        vals  = [
            idx,
            item.get("name") or "-",
            qty,
            f"{int(up):,}" if up is not None else "-",
            f"{int(total):,}" if total is not None else "-",
        ]
        bg = "F2F7FF" if idx % 2 == 0 else "FFFFFF"
        _set_row(ws, r_num, vals, bg=bg, aligns=aligns)
        ws.row_dimensions[r_num].height = 20

    return wb


def build_excel_all(all_data):
    wb = openpyxl.Workbook()

    # ── Sheet 1: 전표 목록 ──
    ws1 = wb.active
    ws1.title = "전표 목록"

    for col, w in zip("ABCDEFG", [14, 24, 16, 10, 14, 8, 16]):
        ws1.column_dimensions[get_column_letter(col_idx := ord(col) - 64)].width = w

    ws1.merge_cells("A1:G1")
    c = ws1["A1"]
    c.value = f"한국도로공사  ·  전표 내역  ({datetime.now().strftime('%Y-%m-%d')} 기준)"
    c.font = _font(bold=True, color="FFFFFF", size=13)
    c.fill = _fill("003087")
    c.alignment = _align()
    ws1.row_dimensions[1].height = 34

    h1 = ["전표번호", "가맹점명", "결제일시", "비용분류", "총금액 (원)", "항목수", "저장일시"]
    a1 = ["center", "left", "center", "center", "right", "center", "center"]
    _set_row(ws1, 2, h1, bold=True, bg="003087", fg="FFFFFF", size=10, aligns=a1)
    ws1.row_dimensions[2].height = 22

    for i, d in enumerate(all_data, 1):
        r   = d["receipt"]
        cnt = len(d["items"])
        vals = [
            r["voucher_no"],
            r["store_name"] or "-",
            r["payment_datetime"] or "-",
            r["category"] or "-",
            f"{int(r['total_amount']):,}" if r["total_amount"] else "-",
            cnt,
            r["created_at"],
        ]
        bg = "F2F7FF" if i % 2 == 0 else "FFFFFF"
        _set_row(ws1, i + 2, vals, bg=bg, aligns=a1)
        ws1.row_dimensions[i + 2].height = 20

    # ── Sheet 2: 품목 상세 ──
    ws2 = wb.create_sheet("품목 상세")

    for col, w in zip("ABCDEF", [14, 24, 28, 10, 14, 14]):
        ws2.column_dimensions[get_column_letter(ord(col) - 64)].width = w

    ws2.merge_cells("A1:F1")
    c2 = ws2["A1"]
    c2.value = "한국도로공사  ·  품목 상세"
    c2.font = _font(bold=True, color="FFFFFF", size=13)
    c2.fill = _fill("003087")
    c2.alignment = _align()
    ws2.row_dimensions[1].height = 34

    h2 = ["전표번호", "가맹점명", "품목명", "수량", "단가 (원)", "금액 (원)"]
    a2 = ["center", "left", "left", "center", "right", "right"]
    _set_row(ws2, 2, h2, bold=True, bg="003087", fg="FFFFFF", size=10, aligns=a2)
    ws2.row_dimensions[2].height = 22

    row_idx = 3
    for band_i, d in enumerate(all_data):
        r = d["receipt"]
        bg = "F2F7FF" if band_i % 2 == 0 else "FFFFFF"
        for item in d["items"]:
            qty   = item.get("quantity") or 1
            up    = item.get("unit_price")
            total = qty * up if up is not None else None
            vals  = [
                r["voucher_no"],
                r["store_name"] or "-",
                item.get("name") or "-",
                qty,
                f"{int(up):,}" if up is not None else "-",
                f"{int(total):,}" if total is not None else "-",
            ]
            _set_row(ws2, row_idx, vals, bg=bg, aligns=a2)
            ws2.row_dimensions[row_idx].height = 20
            row_idx += 1

    return wb


# ── Routes ────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    body = request.get_json(silent=True) or {}
    image_data = body.get("image", "")
    if not image_data:
        return jsonify({"error": "이미지 데이터가 없습니다."}), 400
    try:
        if "," in image_data:
            image_data = image_data.split(",", 1)[1]
        image_bytes = base64.b64decode(image_data)
    except Exception as e:
        return jsonify({"error": f"[decode] {e}"}), 500

    try:
        # MIME 타입 감지
        header = image_bytes[:4]
        if header[:2] == b'\xff\xd8':
            mime = "image/jpeg"
        elif header[:4] == b'\x89PNG':
            mime = "image/png"
        elif header[:4] == b'RIFF':
            mime = "image/webp"
        else:
            mime = "image/jpeg"
        img_part = genai_types.Part.from_bytes(data=image_bytes, mime_type=mime)
    except Exception as e:
        return jsonify({"error": f"[PIL] {e}"}), 500

    try:
        response = client.models.generate_content(model=MODEL, contents=[PROMPT, img_part])
    except Exception as e:
        return jsonify({"error": f"[Gemini] {e}"}), 500

    try:
        text = response.text.strip().lstrip("﻿")
        if "```" in text:
            text = "\n".join(l for l in text.split("\n") if not l.strip().startswith("```")).strip()
        return jsonify(json.loads(text))
    except json.JSONDecodeError as e:
        return jsonify({"error": f"[parse] {e} | raw: {response.text[:200]}"}), 500
    except Exception as e:
        return jsonify({"error": f"[response] {e}"}), 500


@app.route("/save", methods=["POST"])
def save_receipt():
    data = request.get_json()
    voucher_no = next_voucher_no()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO receipts (voucher_no, store_name, payment_datetime,
                                     total_amount, category, filename, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (voucher_no, data.get("store_name"), data.get("payment_datetime"),
             data.get("total_amount"), data.get("category", "기타"),
             data.get("filename"), created_at),
        )
        rid = cur.lastrowid
        for item in data.get("items", []):
            conn.execute(
                "INSERT INTO items (receipt_id, name, quantity, unit_price) VALUES (?,?,?,?)",
                (rid, item.get("name"), item.get("quantity"), item.get("unit_price")),
            )
    return jsonify({"id": rid, "voucher_no": voucher_no})


@app.route("/receipts")
def list_receipts():
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT r.id, r.voucher_no, r.store_name, r.payment_datetime,
                   r.total_amount, r.category, r.filename, r.created_at,
                   (SELECT COUNT(*) FROM items WHERE receipt_id = r.id) AS item_count
            FROM receipts r ORDER BY r.id DESC
        """).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route("/receipts/<int:receipt_id>", methods=["DELETE"])
def delete_receipt(receipt_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM receipts WHERE id = ?", (receipt_id,))
    return jsonify({"ok": True})


@app.route("/export/all")
def export_all():
    with get_conn() as conn:
        receipts = conn.execute("SELECT * FROM receipts ORDER BY id").fetchall()
        all_data = []
        for r in receipts:
            items = conn.execute(
                "SELECT * FROM items WHERE receipt_id = ?", (r["id"],)
            ).fetchall()
            all_data.append({"receipt": dict(r), "items": [dict(i) for i in items]})
    wb  = build_excel_all(all_data)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    today = datetime.now().strftime("%Y%m%d")
    return send_file(buf, as_attachment=True,
                     download_name=f"전표내역_{today}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export/<int:receipt_id>")
def export_one(receipt_id):
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM receipts WHERE id = ?", (receipt_id,)).fetchone()
        if not r:
            return "Not found", 404
        items = conn.execute(
            "SELECT * FROM items WHERE receipt_id = ?", (receipt_id,)
        ).fetchall()
    wb  = build_excel_one(dict(r), [dict(i) for i in items])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"전표_{r['voucher_no']}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    print("서버 시작: http://localhost:5000")
    app.run(debug=True, port=5000)
